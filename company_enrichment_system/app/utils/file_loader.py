from __future__ import annotations

import csv
import io
from collections import OrderedDict

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from app.utils.validators import normalize_company_name

ALLOWED_SUFFIXES = {'.csv', '.xlsx'}


def _deduplicate_preserve_order(rows: list[str]) -> list[str]:
    return list(OrderedDict.fromkeys(rows))


async def load_company_names(file: UploadFile) -> list[str]:
    suffix = None
    if file.filename and '.' in file.filename:
        suffix = file.filename[file.filename.rfind('.'):].lower()

    if suffix not in ALLOWED_SUFFIXES:
        raise HTTPException(status_code=400, detail='Only .csv and .xlsx files are supported')

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail='Uploaded file is empty')

    if suffix == '.csv':
        names = _parse_csv(raw_bytes)
    else:
        names = _parse_xlsx(raw_bytes)

    cleaned = [normalize_company_name(name) for name in names if name and normalize_company_name(name)]
    deduped = _deduplicate_preserve_order(cleaned)

    if not deduped:
        raise HTTPException(status_code=400, detail='No valid company names found in first column')

    return deduped


def _parse_csv(raw_bytes: bytes) -> list[str]:
    decoded = raw_bytes.decode('utf-8-sig', errors='ignore')
    reader = csv.reader(io.StringIO(decoded))
    return [row[0] for row in reader if row and row[0] and row[0].strip()]


def _parse_xlsx(raw_bytes: bytes) -> list[str]:
    buffer = io.BytesIO(raw_bytes)
    workbook = load_workbook(buffer, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[str] = []

    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        if not row:
            continue
        value = row[0]
        if value is None:
            continue
        value_str = str(value).strip()
        if value_str:
            rows.append(value_str)

    workbook.close()
    return rows
